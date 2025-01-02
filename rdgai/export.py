from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
import pandas as pd

from .apparatus import Doc


def export_variants_to_excel(doc:Doc, output:Path):
    """ Export the variants to an Excel file."""
    wb = Workbook()
    header_font = Font(bold=True)

    relation_types = doc.relation_types

    # Rename the default sheet
    ws = wb.active
    ws.title = 'Variants'

    headers = [
        'App ID', 'Context', 
        'Active Reading ID', 'Passive Reading ID',
        'Active Reading Text', 'Passive Reading Text', 
        'Description', 'Relation Type(s)',
    ]

    for col_num, header in enumerate(headers, start=1):  # Start from column A (1)
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font

    current_row = 2
    for app in doc.apps:
        for pair in app.non_redundant_pairs:
            ws[f'A{current_row}'] = str(app)
            ws[f'B{current_row}'] = app.text_in_context()
            ws[f'C{current_row}'] = pair.active.n
            ws[f'D{current_row}'] = pair.passive.n
            ws[f'E{current_row}'] = pair.active.text
            ws[f'F{current_row}'] = pair.passive.text
            ws[f'G{current_row}'] = pair.get_description()
            
            for relation_type_index, relation_type in enumerate(pair.types):
                column = ord('H') + relation_type_index
                ws[f'{chr(column)}{current_row}'] = str(relation_type)
            
            current_row += 1

    data_val = DataValidation(type="list",formula1=f'"{",".join(relation_types.keys())}"')
    ws.add_data_validation(data_val)

    max_relation_types = max(10, max(len(pair.types) for app in doc.apps for pair in app.pairs))
    end_column = chr(ord('H') + max_relation_types - 1)

    data_val.add(f"H2:{end_column}{current_row}")

    # Create new sheet with descriptions of categories and counts
    categories_worksheet = wb.create_sheet('Categories')

    # Add a header to the "Category" column
    headers = ['Category', 'Inverse', 'Count', 'Inverse Count', 'Total', 'Description']
    for col_num, header in enumerate(headers, start=1): 
        cell = categories_worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font

    # Populate the categories from relation_types.keys()
    for idx, category in enumerate(relation_types.values(), start=2):  # Start from row 2
        category_name = str(category)
        inverse_name = str(category.inverse) if category.inverse else category_name
        categories_worksheet[f'A{idx}'] = category_name
        categories_worksheet[f'B{idx}'] = inverse_name
        categories_worksheet[f'C{idx}'] = f'=COUNTIF(Variants!G:{end_column}, "{category_name}")'
        categories_worksheet[f'D{idx}'] = f'=COUNTIF(Variants!G:{end_column}, "{inverse_name}")'
        categories_worksheet[f'E{idx}'] = f'=SUM(C{idx}:D{idx})'
        categories_worksheet[f'F{idx}'] = category.description

    wb.save(output)


def import_classifications_from_dataframe(doc:Doc, variants_df:pd.DataFrame, output:Path, responsible:str|None=None):
    variants_df.fillna('', inplace=True)
    apps_dict = {str(app): app for app in doc.apps}
    relation_types = doc.relation_types
    for _, row in variants_df.iterrows():
        app_id = row['App ID']
        active_reading_id = row['Active Reading ID']
        passive_reading_id = row['Passive Reading ID']
        description = row['Description'].strip() if 'Description' in row else None
        app = apps_dict[app_id]

        types = set(row[key] for key in row.keys() if (key.startswith('Relation Type') or key.startswith('Unnamed: ')) and row[key])
        # for type in types:
        #     assert type in relation_types, f'{type} not in {relation_types.keys()}'
        types = set(relation_types[type] for type in types if type in relation_types)

        for pair in app.pairs:
            if str(pair.active.n) == str(active_reading_id) and str(pair.passive.n) == str(passive_reading_id):
                # Add relations
                for type in types - pair.types:
                    pair.add_type_with_inverse(type, responsible=responsible)
            
                # Remove relations
                for type in pair.types - types:
                    pair.remove_type_with_inverse(type)

                if description:
                    pair.add_description(description)
                elif description == "":
                    # remove description if it is an empty string
                    # don't do anything if description is 'None'
                    pair.remove_description()

    doc.write(output)        
